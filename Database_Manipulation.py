"""Imported libraries uses will be listed below:
    xlrd - opening and reading excel workbooks
    xlwt - writing excel workbooks
    xlutils - copying excel documents
    pdfminer - extracting data from pdfs
    csv - manipulating and working with csv documents
    shutil - 
    os - having operating software functions
    cstringIO - stringIO objects
    forex_python.converter - exchange rate converter
    Google_Drive_Manipulation - using the Google Dirve and Sheets API"""

import xlrd, xlwt, pdfminer, csv, shutil, os, xlutils, sys, openpyxl, gspread
# from cstringIO import stringIO
from CurrencyConverterNew import *
from decimal import *
from Google_API_Manipulation import *
from datetime import *
from xlutils.copy import copy
from openpyxl import styles
from write_log import *
#from gspread import *

reload(sys)
sys.setdefaultencoding('utf-8')

""" Currency Rate List defined here, and called so that it is only called once per program iteration"""
global currency_dictionary, currency_list

currency_dictionary =  ({'USD': ['Rate', 'Price', 'New Price', 'New Price (USD)', 'Rate - USD']},
                        {'EUR': ['New Price(Euro)', 'Price Euro', 'New Price EUR', 'New Price (EUR)', 'Price \nEUR/SMS', 'Price in EUR']},
                        {'GBP': ['Price in GBP']},
                        {'CNY': []},
                        {'MXN': []},
                        {'AUD': ['Price in AUD']},
                        {'GW': ['GW0', 'GW111']})

# """Support only exists for USD, EUR right now, need to define dictionary for others"""
currency_list = ['USD', 'EUR', 'GBP', 'CNY', 'MXN', 'AUD', 'GW']

#value for date to be entered
global today, tomorrow
tomo = date.today() + timedelta(days = 1)
tomorrow = str(tomo)[-5:]
today = str(date.today())[-5:]

"""BST class is used for building binary search tree datastructure, used to pull from database
    and track data changes"""
class bst():
    class node():
        def __init__(self, hashkey, val):
            self.l_child = None
            self.r_child = None
            self.key = hashkey
            self.data = val

    def size(self, root):
        count = 1
        if root == None:
            return 0
        else:
            count = count + self.size(root.l_child)
            count = count + self.size(root.r_child)
            return count

    def changes_to_email(self, notify_list, email, column_title):
        book = openpyxl.Workbook()
        sheet = book.active
        
        for y in range(len(column_title) - 1):
            value = column_title[y+1]
            sheet.cell(row=1,column=y+1).value = value
        
        for i in range(len(notify_list)):
            temp = notify_list.pop(0)
            for j in range(len(temp)):
                if j == 0:
                    pass
                else:
                    value = temp[j]
                    sheet.cell(row=i+2,column=j).value = value
        book.save('Changes.xlsx')
        

    def database_build(self, root, edate, change_root, wholesale_root):
        filename = 'Rates for ' + str(edate)
        # """Attempts to locate file using the filename in the 'Compiled Data Folder' """"
        # file_id = find_file_id_using_parent(filename, '0BzlU44AWMToxYmdRR1hHVXJiQ1E')
        day_before = edate
        days = 0
        book_found = False
        filename_old = filename

        while(days <= 10):
            if os.path.isfile(filename_old + '.xls'):
                log_file().info("Source Compiler.log", "File found with filename %s." %filename_old)
                print ("File found with filename %s." %filename_old)
                book_found = True
                break
            else:
                log_file().info("Source Compiler.log", "No file found for %s." %filename_old)
                print ("No file found for %s." %filename_old)
                day_before = day_before - timedelta(days = 1)
                filename_old = 'Rates for ' + str(day_before)
                filename_old = 'Rates for ' + str(day_before)
                days = days + 1

        # If a previous file has not been found, it will generate a worksheet.
        if not book_found:
            #if not os.path.isfile(filename_old + '.xlsx'):
            log_file().info("Source Compiler.log", "New Rates sheet created.  Either no previous versions or most recent version is more than 10 days old.")
            print ("New Rates sheet created.  Either no previous versions or most recent version is more than 10 days old.")
            book = xlwt.Workbook(style_compression=2)
            sheet = book.add_sheet("Sheet", cell_overwrite_ok=True)
            filename_old = filename
            book.save(filename_old + '.xls')
            day_before = edate
            return

        currency_rate = get_rate_for_date(day_before)
        filename = filename_old + '.xls'
        book = xlrd.open_workbook(filename, formatting_info=True)
        sheet = book.sheet_by_index(0)
        rownum = sheet.nrows
        colnum = 11
        for i in range(rownum-1):
            i = i + 1
            if sheet.cell(i, 1).value == None:
                break
            string = ''
            """provider = [hash key, country, network, mcc, mnc, mccmnc, rates, curr, converted rate, source, date, change]"""
            provider = [0]
            for j in range(colnum):
                #Catches weird country namings
                if j == 0:
                    value = sheet.cell(i,j).value
                    value = value.lower()
                    value = value.title()
                    provider.append(value)
                elif j == 7:
                    if provider[7] == 'CURR':
                        provider.append(sheet.cell(i, j).value)
                    elif not provider[7] == 'USD':
                        curr = 0
                        for x in range(len(currency_list)):
                            if provider[7] in currency_list[x]:
                                curr = x
                                break

                        converted = currency_rate[curr]*float(provider[6])
                        provider.append(converted)
                    else:
                        provider.append(sheet.cell(i,j).value)
                elif j == 9:
                    provider.append(convert_date(sheet.cell(i, j).value))
                else:
                    provider.append(sheet.cell(i, j).value)
                if j < 5:
                    string = string + str(sheet.cell(i, j).value).encode("utf-8")
                else:
                    pass

            if provider[11] == '':
                provider[11] == '-----'

            string = string + str(provider[9]).decode('utf-8')
            provider[0] = hash(string)
            #provider.append(0)
            if provider[10] >= edate:
                xfx = sheet.cell_xf_index(i, 7)
                xf = book.xf_list[xfx]
                bgx = xf.background.pattern_colour_index
                if bgx == 10:
                    provider[11] = "Increase"
                elif bgx == 17:
                    provider[11] = "Decrease"
                else:
                    provider[11] = "-----"
            else:
                provider[11] = "-----"

            self.insert(root, self.node(provider[0], provider), change_root)
            
        w_sheet = book.sheet_by_index(1)
        rownum = w_sheet.nrows
        colnum = 11
        for i in range(rownum-1):
            i = i + 1
            if w_sheet.cell(i, 1).value == None:
                break
            string = ''
            """provider = [hash key, country, network, mcc, mnc, mccmnc, rates, curr, converted rate, source, date, change]"""
            provider = [0]
            for j in range(colnum):
                if j == 7:
                    if provider[7] == 'CURR':
                        provider.append(w_sheet.cell(i, j).value)
                    elif not provider[7] == 'USD':
                        curr = 0
                        for x in range(len(currency_list)):
                            if provider[7] in currency_list[x]:
                                curr = x
                                break

                        converted = currency_rate[curr]*float(provider[6])
                        provider.append(converted)
                    else:
                        provider.append(w_sheet.cell(i,j).value)
                elif j == 9:
                    provider.append(convert_date(w_sheet.cell(i, j).value))
                else:
                    provider.append(w_sheet.cell(i, j).value)
                if j < 5:
                    string = string + str(w_sheet.cell(i, j).value).encode("utf-8")
                else:
                    pass

            if provider[11] == '':
                provider[11] == '-----'

            string = string + str(provider[9]).decode('utf-8')
            provider[0] = hash(string)
            #provider.append(0)
            if provider[10] >= edate:
                xfx = w_sheet.cell_xf_index(i, 7)
                xf = book.xf_list[xfx]
                bgx = xf.background.pattern_colour_index
                if bgx == 10:
                    provider[11] = "Increase"
                elif bgx == 17:
                    provider[11] = "Decrease"
                else:
                    provider[11] = "-----"
            else:
                provider[11] = "-----"

            self.insert(wholesale_root, self.node(provider[0], provider), change_root)


    def insert(self, root, node, change_root):
        if root is None:
            root = node
        else:
            """if statements are based on hash key of the strings built"""
            if root.key > node.key:
                if root.l_child is None:
                    root.l_child = node
                else:
                    self.insert(root.l_child, node, change_root)
            elif root.key < node.key:
                if root.r_child is None:
                    root.r_child = node
                else:
                    self.insert(root.r_child, node, change_root)
            elif root.key == node.key:
                if root.data[8] == 'Converted Rate':
                    pass
                # """Comparing rates of various nodes, typecasting to float"""
                # """Price decreased"""
                elif float(root.data[6]) > float(node.data[6]):
                    node.data[11] = 'Decrease'
                    root.data = node.data
                    data = root.data
                    key = root.key                    
                    if node.data[10] >= date.today():
                        new_node = self.node(key, data)
                        temp = new_node.data
                        new = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
                        new[11] = temp[9]
                        new[10] = 0.0
                        new[9] = 0.0
                        new[8] = temp[6]
                        new[7] = temp[5]                    
                        #if temp[4] < 10:
                            #new[6] = '0' + str(int(temp[4]))
                        #else:
                            #new[6] = str(int(temp[4]))
                        new[6] = temp[4]
                        new[5] = temp[3]
                        new[4] = temp[2]
                        new[3] = temp[1]
                        new[2] = 'CC'        
                        new[1] = 'Region'
                        new[0] = temp[0]
                        new_node.data = new                    
                        self.insert_new(change_root, new_node)                        

                # """Price increased"""
                elif float(root.data[6]) < float(node.data[6]):
                    node.data[11] = 'Increase'
                    root.data = node.data
                    data = root.data
                    key = root.key
                    if node.data[10] >= date.today():
                        new_node = self.node(key, data)
                        temp = new_node.data
                        new = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
                        new[11] = temp[9]
                        new[10] = 0.0
                        new[9] = 0.0
                        new[8] = temp[6]
                        new[7] = temp[5]                    
                        #if temp[4] < 10:
                            #new[6] = '0' + str(int(temp[4]))
                        #else:
                            #new[6] = str(int(temp[4]))
                        new[6] = temp[4]
                        new[5] = temp[3]
                        new[4] = temp[2]
                        new[3] = temp[1]
                        new[2] = 'CC'        
                        new[1] = 'Region'
                        new[0] = temp[0]
                        new_node.data = new                    
                        self.insert_new(change_root, new_node)
                # """no change"""
                else:
                    pass

    def insert_new(self, root, node):
        """provider = [hash key, country, network, mcc, mnc, mccmnc, rates, curr, converted rate, source, date, change]"""
        #[hash, region, cc, country, network, mcc, mnc, mccmnc, cost, price, profit margin, source]
        if root is None:
            root = node
        else:
            if root.key > node.key:
                if root.l_child is None:
                    root.l_child = node
                else:
                    self.insert_new(root.l_child, node)
            elif root.key < node.key:
                if root.r_child is None:
                    root.r_child = node
                else:
                    self.insert_new(root.r_child, node)

    def insert_price(self, root, node, notify_list):
        if root is None:
            root = node
        else:
            """if statements are based on hash key of the strings built"""
            if root.key > node.key:
                if root.l_child is None:
                    root.l_child = node
                else:
                    self.insert_price(root.l_child, node, notify_list)
            elif root.key < node.key:
                if root.r_child is None:
                    root.r_child = node
                else:
                    self.insert_price(root.r_child, node, notify_list)
            elif root.key == node.key:
                # provider list = [hash, region, cc, country, network, mcc, mnc, mccmnc, cost, price, profit margin, source]
                root.data[1] = node.data[1]
                root.data[2] = node.data[2]
                
                root.data[9] = node.data[9]
                profit = (root.data[9] - root.data[8]) / root.data[8]
                root.data[10] = profit
                if profit < 0.2:
                    notify_list.append(root.data)
    
    def in_order_print(self, root):
        if not root:
            return
        self.in_order_print(root.l_child)
        log_file().info("Source Compiler.log", "In order print: " + str(''.join(str(root.data))))
        print root.data
        self.in_order_print(root.r_child)

    def pre_order_print(self, root):
        if not root:
            return
        log_file().info("Source Compiler.log", "Pre order print: " + str(''.join(str(root.data))))        
        print root.data
        self.pre_order_print(root.l_child)
        self.pre_order_print(root.r_child)

    def price_build(self, root, filename):
        provider_list = ['Mitto AG', 'Tata Communications', 'CLX Networks', 'Tedexis', 'UPM Telecom']
        provider_dictionary = ({'Mitto AG': ['Mitto', 'Mitto Wholesale']},
                               {'Tata Communications': ['TATA']},
                               {'CLX Networks': ['CLX']},
                               {'Tedexis': ['Tedexis']},
                               {'UPM Telecom': ['UPM']})
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(4)
        rownum = sheet.nrows
        colnum = sheet.ncols
        notify_list = []
        for i in range(rownum):
            # provider list = [hash, region, cc, country, network, mcc, mnc, mccmnc, cost, price, profit margin, source]
            provider = [0]
            no_cost = False
            string = ''
            if i == 0:
                pass
            else:
                for j in range(colnum):
                    # Splits MCCMNC to MCC and MNC, and appends all 3 to provider
                    if j == 4:
                        value = sheet.cell(i,j).value.decode('utf-8')
                        mcc = value[:3]
                        mnc = value[3:5]
                        provider.append(mcc)
                        provider.append(mnc)
                        provider.append(value)
                        string = string + mcc + mnc + value
                    # appends Source - catches unusual naming
                    elif j == 8:
                        value = sheet.cell(i,j).value.decode('utf-8')
                        for k in range(len(provider_list)-1):
                            if value in provider_dictionary[k][provider_list[k]]:
                                value = provider_list[k].decode('utf-8')
                                break
                        provider.append(value)               
                        string = string + value    
                    # appends Region | CC | Country | Network | Cost | Price | Profit Margin
                    else:
                        value = sheet.cell(i,j).value
                        #hash for country and network, decodes to unicode
                        if j == 2 or j == 3:
                            value = value.decode('utf-8')
                            string = string + value
                        #converts cost and price to float values
                        elif j == 5 or j == 6:
                            try:
                                value = float(value)
                            except ValueError:
                                no_cost = True
                        
                        provider.append(value) 
                
                #String to hash = Country, Network, MCC, MNC, MCCMNC, Source
                provider[0] = hash(string)
                
                #if no_cost:
                    #notify_list.append(provider)
                #else:
                self.insert_price(root, self.node(provider[0], provider), notify_list)

        return notify_list

    # """Builds BST structure for all sources in filename that is taken in.  Structure built off of 
    #     root taken in as argument"""
    def source_build(self, root, filename, change_root):
        try:
            book = xlrd.open_workbook(filename, 'rb')
            sheet = book.sheet_by_index(0)
            rownum = sheet.nrows #should be 10
            colnum = sheet.ncols
            for i in range(rownum-1):
                i = i + 1
                string = ''
                """provider = [hash key, country, network, mcc, mnc, mccmnc, rates, curr, converted rate, source, date, change]"""
                provider = [0]
                for j in range(colnum):
                    provider.append(sheet.cell(i,j).value) 
                    if j < 5:
                        value = sheet.cell(i,j).value
                        #if j == 2 or j == 3:
                            #if value < 10:
                                #value = '0' + str(int(value))
                            #else:
                                #value = str(int(value))
                        string = string + str(value).decode("utf-8")
                    else:
                        pass
                
                string = string + str(provider[9]).decode('utf-8')
                provider[0] = hash(string)
                provider[10] = convert_date(provider[10])
                provider.append('-----')
                
                self.insert(root, self.node(provider[0], provider), change_root)
            
        except:
            error = sys.exc_info()[0]
            log_file().error("Source Compiler.log", "Error in bst().source_build: %s" % error)
            print("Error in bst().source_build: %s" % error)

    """Takes in node, and list.  Builds a pre-order list of node.data and stores in list taken in"""
    def to_database(self, root, templist):
        if not root:
            return
        templist.append(root.data)
        self.to_database(root.l_child, templist)
        self.to_database(root.r_child, templist)

    def write(self, root, edate, wholesale_root):
        try:
            book = xlwt.Workbook(style_compression=2)
            sheet = book.add_sheet("Premium",cell_overwrite_ok=True)
            w_sheet = book.add_sheet("Wholesale",cell_overwrite_ok=True)
            filename = 'Rates for ' + str(edate) + '.xls'
            final_list = []
            length = 11 #lenght of provider list - 2 (hash key and change value)
    
            self.to_database(root, final_list)
            title = []
            title.append(final_list.pop(0))
            temp = format().quicksort(final_list, 1)
            final_list = title + temp
    
    
            # final_list = title + final_list
            count = 0
            for x in range(len(final_list)):
                provider = final_list.pop(0)
                if provider[10] != 'Effective Date':
                    if provider[10] < edate:
                        provider[11] = "-----"
                if provider[8] == 0:
                    count = count + 1
                    pass
                # print len(final_list)
                else:
                    for k in range(length):
                        if x == 0:
                            st = xlwt.easyxf('align: horiz center')
                            sheet.write(x - count,k,provider[k+1],st)
                        else:
                            if k == 5:
                                st = xlwt.easyxf('align: horiz right')
                                sheet.write(x - count,k,provider[k+1],st)
                            elif k == 7:
                                # price increased
                                if provider[11] == 'Increase':
                                    st = xlwt.easyxf('pattern: pattern solid, fore_color red; align: horiz right')
                                    sheet.write(x - count,k,float(provider[k+1]),st)
                                    # print "marker 1"
                                # price decreased
                                elif provider[11] == 'Decrease':
                                    st = xlwt.easyxf('pattern: pattern solid, fore_color green; align: horiz right')
                                    sheet.write(x - count,k,float(provider[k+1]),st)
                                    # print "marker 2"
                                else:
                                    st = xlwt.easyxf('align: horiz right')
                                    sheet.write(x - count,k,provider[k+1],st)
                                    # print "marker 3"
                            elif k == 9:
                                sheet.write(x - count,k,str(provider[k+1]))
                            else:
                                st = xlwt.easyxf('align: horiz left')
                                sheet.write(x - count,k,provider[k+1],st)
                                # print "marker 4"
                            
            sheet.col(0).width = 6500
            sheet.col(1).width = 8000
            sheet.col(2).width = 2500
            sheet.col(6).width = 2500 
            sheet.col(8).width = 5000
            sheet.set_panes_frozen(True)
            sheet.set_horz_split_pos(1)
            
            final_list = []
            length = 11 #lenght of provider list - 2 (hash key and change value)
    
            self.to_database(wholesale_root, final_list)
            title = []
            title.append(final_list.pop(0))
            temp = format().quicksort(final_list, 1)
            final_list = title + temp
    
            for x in range(len(final_list)):
                provider = final_list.pop(0)
                if provider[10] != 'Effective Date':
                    if provider[10] < edate:
                        provider[11] = "-----"
                # print len(final_list)
                for k in range(length):
                    if x == 0:
                        st = xlwt.easyxf('align: horiz center')
                        w_sheet.write(x,k,provider[k+1],st)
                    else:
                        if k == 5:
                            st = xlwt.easyxf('align: horiz right')
                            w_sheet.write(x,k,provider[k+1],st)
                        elif k == 7:
                            # price increased
                            if provider[11] == 'Increase':
                                st = xlwt.easyxf('pattern: pattern solid, fore_color red; align: horiz right')
                                w_sheet.write(x,k,float(provider[k+1]),st)
                                # print "marker 1"
                            # price decreased
                            elif provider[11] == 'Decrease':
                                st = xlwt.easyxf('pattern: pattern solid, fore_color green; align: horiz right')
                                w_sheet.write(x,k,float(provider[k+1]),st)
                                # print "marker 2"
                            else:
                                st = xlwt.easyxf('align: horiz right')
                                w_sheet.write(x,k,provider[k+1],st)
                                # print "marker 3"
                        elif k == 9:
                            w_sheet.write(x,k,str(provider[k+1]))
                        else:
                            st = xlwt.easyxf('align: horiz left')
                            w_sheet.write(x,k,provider[k+1],st)
                            # print "marker 4"        
            w_sheet.col(0).width = 6500
            w_sheet.col(1).width = 8000
            w_sheet.col(2).width = 2500
            w_sheet.col(6).width = 2500 
            w_sheet.col(8).width = 5000
            w_sheet.set_panes_frozen(True)
            w_sheet.set_horz_split_pos(1)
    
            log_file().info("Source Compiler.log", 'Successfully written. Data for %s is now queued to upload.' %str(edate))
            print ('Successfully written. Data for %s is now queued to upload.' %str(edate))
            #clear out previous working versions
            
            #Production version
            book.save(filename)
            file_id = find_file_id_using_parent(filename, '0BzlU44AWMToxYmdRR1hHVXJiQ1E')
            if file_id != None:
                delete_file(file_id)
            upload_excel(filename)
            move_to_folder_using_name(filename, '0BzlU44AWMToxYmdRR1hHVXJiQ1E')
            temp_file_id = find_file_id_using_parent('Rates for ' + str(edate), '0BzlU44AWMToxNEtxSWROcjkzYVE')
            if temp_file_id != None:
                delete_file(temp_file_id)        
            upload_as_gsheet(filename, 'Rates for ' + str(edate))
            move_to_folder_using_name('Rates for ' + str(edate), '0BzlU44AWMToxNEtxSWROcjkzYVE')
            
            # Development version uses test folders
            #book.save(filename)
            #file_id = find_file_id_using_parent(filename, '0BzlU44AWMToxSTNfYTFkdm5MZEE')
            #if file_id != None:
                #delete_file(file_id)
            #upload_excel(filename)
            #move_to_folder_using_name(filename, '0BzlU44AWMToxSTNfYTFkdm5MZEE')
            #temp_file_id = find_file_id_using_parent('Test Rates for ' + str(edate), '0BzlU44AWMToxYW5iWmFWVWdzNnM')
            #if temp_file_id != None:
                #delete_file(temp_file_id)        
            #upload_as_gsheet(filename, 'Test Rates for ' + str(edate))
            #move_to_folder_using_name('Test Rates for ' + str(edate), '0BzlU44AWMToxYW5iWmFWVWdzNnM')      
        
        except:
            error = sys.exc_info()[0]
            log_file().error("Source Compiler.log", "Error in bst().write: %s" % error)              
            print("Error in bst().write: %s" % error)            

    def write_price(self, change_root, wholesale_root):
        #book = xlwt.Workbook()
        #sheet = book.add_sheet('Sheet', cell_overwrite_ok=True)
        
        #rb = xlrd.open_workbook('Copy of Hook Full Price List interns.xlsx')
        #r_sheet = rb.sheet_by_index(4)
        #book = copy(rb)
        #sheet = book.get_sheet(4)
        
        book = openpyxl.load_workbook('Hook Full Price List interns.xlsx')
        #r_sheet = rb['A-Z INTERNAL']
        #book = copy(rb)
        sheet = book['A-Z INTERNAL']
        
        final_list = []
        self.to_database(change_root, final_list)
        title = []
        title.append(final_list.pop(0))
        temp = format().quicksort(final_list, 1)
        final_list = title + temp
        colnum = 9
        
        for i in range(len(final_list)):
            provider = final_list.pop(0)
            provider.pop(5)
            provider.pop(5)
            for j in range(colnum):
                if i == 0:
                    #st = xlwt.easyxf('align: horiz center')
                    row = sheet.row_dimensions[1]
                    row.alignment = styles.Alignment(horizontal='center')
                    
                    #for row in rows:
                        #cell = row[i][j]
                        #cell.alignment = Alignment(horizontal='center')                    
                            
                    sheet.cell(row=i+1,column=j+1).value = provider[j+1]
                else:
                    sheet.cell(row=i+1,column=j+1).value = provider[j+1]
        
        sheet.freeze_panes = sheet['A2']
        
        book.save('Pricing Sheet.xlsx')
        log_file().info("Source Compiler.log", 'Pricing sheet has been updated.')
        print ('Pricing sheet has been updated.')
        #upload_excel('Pricing Sheet.xlsx')
        
"""Convert class performs all file conversions"""
class convert():
    """CSV_TO_EXCEL takes in a string argument of the filename, and returns
        string with filename of converted document, removes original document"""
    def csv_to_excel(self, filename):
        file = open(filename, 'rb')
        read = csv.reader((file), delimiter = ',')
        book = xlwt.Workbook()
        sheet = book.add_sheet("Sheet 1")

        for rowi, row in enumerate(read):
            for coli, value in enumerate(row):
                sheet.write(rowi,coli,value)

        index = filename.rfind('.')
        filename1 = filename[:index]
        filename1 = filename1 + '.xls'
        book.save(filename1)
        file.close()
        # os.remove(filename)
        return filename1    




    """ EXCEL_TO_CSV takes in a string argument of the filename, and returns
        string with filename of converted document, removes original document"""
    def excel_to_csv(self, filename):
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        index = filename.rfind('.')
        filename1 = filename[:index] 
        filename1 = filename1 + '.csv'
        csvfile = open(filename1, 'wb')
        write = csv.writer(csvfile, quoting=csv.QUOTE_ALL)

        for row in range(sheet.nrows):
            write.writerow(sheet.row_values(row))

        csvfile.close()
        # os.remove(filename)
        return filename1

    """ EXCEL_TSV_TO_CSV takes in a string argument of the filename, and returns
        string with filename of converted document, removes original document.  Converts
        excel files found in tsv format to csv"""
    def excel_tsv_to_csv(self, filename):
        tsvfile = open(filename, 'rb')
        read = csv.reader(tsvfile, dialect = csv.excel_tab)
        index = filename.rfind('.')
        filename1 = filename[:index]
        filename1 = filename1 + '.csv'
        csvfile = open(filename1, 'wb')
        write = csv.writer(csvfile, dialect = csv.excel)

        for row in read:
            write.writerow(row)

        tsvfile.close()
        csvfile.close()
        # os.remove(filename)
        return filename1

    """ PDF_TO_CSV takes in a string argument of filename, and returns string with 
        filename to converted document, removes original document."""
    # def pdf_to_csv(self, filename):
        # reference Email_Parser.py for this section


"""Format class contains all formatting methods that will be used"""
class format():
    global column_dictionary, column_list
    column_dictionary =({'Country': ['Country', 'CountryName']},
                        {'Network': ['Network', 'OperatorName', 'Operator']},
                        {'Country/Network': ['Country/Operator', 'Region/Operator', 'Country/Network']},
                        {'MCC': ['MCC']},
                        {'MNC': ['MNC']},
                        {'MCCMNC': ['MCCMNC', 'Network code', 'IMSI', 'MCC MNC']},
                        ({'Rate': ['Rate', 'Price', 'New Price', 'New Price(Euro)', 'Price Euro', 'New Price EUR', 
                            'New Price (EUR)', 'Price \nEUR/SMS', 'New Price (USD)', 'Rate - USD', 'Price in GBP',
                            'Price in AUD', 'Price in EUR', 'GW0', 'GW111']}))

    column_list = ['Country', 'Network', 'Country/Network', 'MCC', 'MNC', 'MCCMNC', 'Rate'] # , 'CURR', 'Source']

    def calltrade(self, filename, edate):
        currency_rate = get_rate_for_date(edate)
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        rownum = sheet.nrows
        colnum = sheet.ncols
        new_book = xlwt.Workbook()
        sheet_wr = new_book.add_sheet("Sheet", cell_overwrite_ok=True)
        for i in range(rownum):
            for j in range(colnum):
                value = sheet.cell(i,j).value
                if i == 0:
                    sheet_wr.write(i,j,value)
                else:
                    if j == 6:
                        sheet_wr.write(i,j,'EUR')
                    elif j == 7:
                        val = float(value)*currency_rate[1]
                        sheet_wr.write(i,j,val)
                    else:
                        sheet_wr.write(i,j,value)
        new_book.save(filename)    
        
    # """ excel_filter takes and removes empty rows from a FORMATTED document """
    def excel_filter(self, filename):
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        new_book = xlwt.Workbook()
        sheet_wr = new_book.add_sheet("sheet", cell_overwrite_ok = True)

        for i in range(sheet.nrows):
            if sheet.cell(i,0).value == '':
                break
            for j in range(sheet.ncols):
                value = sheet.cell(i,j).value
                sheet_wr.write(i,j,value)

        index = filename.rfind('.')
        filename1 = filename[:index]
        filename1 = filename1 + ' and FILTERED.xls'
        new_book.save(filename1)
        return filename1
        # os.remove(filename)


    # """ EXCEL_FORM takes in both .xls or .xlsx and rearranges the columns to be
    #   correctly ordered.  takes in filename as string, returns new filename."""
    def excel_format(self, filename, source, sheetindex, edate):
        try:
            currency_rate = get_rate_for_date(edate)
            book = xlrd.open_workbook(filename)
            sheet = book.sheet_by_index(sheetindex)
            new_book = xlwt.Workbook()
            sheet_wr = new_book.add_sheet("sheet", cell_overwrite_ok = True) 
    
            """Forcing the header of the excel format"""
            sheet_wr.write(0,0, 'Country')
            sheet_wr.write(0,1, 'Network')
            sheet_wr.write(0,2, 'MCC')
            sheet_wr.write(0,3, 'MNC')
            sheet_wr.write(0,4, 'MCCMNC')
            sheet_wr.write(0,5, 'Rate')
            sheet_wr.write(0,6, 'CURR')
            # sheet_wr.write(0,7, 'Source')
            sheet_wr.write(0,7, 'Converted Rate')
            sheet_wr.write(0,8, 'Source')
            sheet_wr.write(0,9, 'Effective Date')
            """Freezing Header Row"""
            sheet_wr.set_panes_frozen(True)
            sheet_wr.set_horz_split_pos(1)
    
            rownum = sheet.nrows
            colnum = sheet.ncols
    
            edate_effective = edate + timedelta(days = 1)
            rate_present = False
            mccmnc_absent=True
            mcc_absent=True
            mnc_absent=True
            mnc_val=[0]
            mcc_val=[0]
            mccmnc_val=[0]
    
    
            for row in range(rownum):
                for y in range(colnum):
                    # """Country"""
                    if sheet.cell(row,y).value in column_dictionary[0][column_list[0]]:
                        for x in range(row+1, rownum):
                            value = sheet.cell(x,y).value
                            if value == '':
                                pass
                            else:
                                value = value.lower()
                                value = value.title()
                                sheet_wr.write(x-row,0,value)
                    # """Network"""        
                    elif sheet.cell(row,y).value in column_dictionary[1][column_list[1]]:
                        for x in range(row+1, rownum):
                            value = sheet.cell(x,y).value
                            sheet_wr.write(x-row,1,value)
                    # """Country/Network"""
                    elif sheet.cell(row,y).value in column_dictionary[2][column_list[2]]:
                        for x in range(row+1, rownum):
                            value = self.separator(sheet.cell(x,y).value)
                            sheet_wr.write(x-row,0,value[0])
                            sheet_wr.write(x-row,1,value[1])
                    # """MCC"""
                    elif sheet.cell(row,y).value in column_dictionary[3][column_list[3]]:
                        mcc_absent=False
                        for x in range(row+1, rownum):
                            value = sheet.cell(x,y).value
                            mcc_val.append(value)
                            sheet_wr.write(x-row,2,value)
                    # """MNC"""
                    elif sheet.cell(row,y).value in column_dictionary[4][column_list[4]]:
                        mnc_absent=False
                        for x in range(row+1, rownum):
                            value = sheet.cell(x,y).value
                            mnc_val.append(value)
                            sheet_wr.write(x-row,3,value)
                    # """MCCMNC"""
                    elif sheet.cell(row,y).value in column_dictionary[5][column_list[5]]:
                        mccmnc_absent=False
                        for x in range(row+1, rownum):
                            value = sheet.cell(x,y).value
                            mccmnc_val.append(value)
                            sheet_wr.write(x-row,4,value)
    
                    # """Rate"""
                    elif sheet.cell(row,y).value in column_dictionary[6][column_list[6]]:
                        rate_present = True
                        for x in range(len(currency_list)):
                            if sheet.cell(row,y).value in currency_dictionary[x][currency_list[x]]:
                                i = x
                                break
                        for x in range(row+1, rownum):
                            if sheet.cell(x,y).value == '-' or sheet.cell(x,y).value == '':
                                value = 0
                            elif str(sheet.cell(x, y).value)[0] == '$':
                                temp = sheet.cell(x, y).value
                                value = temp[2:]
                                value = float(value)
                            else:
                                value = sheet.cell(x,y).value
                            if sheet.cell(x,y).value == '':
                                pass
                            else:
                                sheet_wr.write(x-row,5,value)
                                if currency_list[i] == 'GW':
                                    currency = 'USD'      
                                    # """Adjust converted value - for GW0 and GW111"""
                                    converted = float(str(value)[-4:])/10000
                                    sheet_wr.write(x - row, 5, converted)
                                elif not currency_list[i] == 'USD':
                                    currency = currency_list[i]
                                    converted = currency_rate[i]*float(value)
                                else:
                                    currency = currency_list[i]
                                    converted = value
                                sheet_wr.write(x-row,6,currency)
                                sheet_wr.write(x-row,7,converted)
                                sheet_wr.write(x-row,8,source)
                                sheet_wr.write(x-row,9,str(edate_effective))                                
                    else:
                        pass
                    
            # """ Computing missing MNC, MCC or MCCMNC Values"""
            
            # # """MCCMNC is absent"""
            if mcc_absent == False and mnc_absent==False and mccmnc_absent==True:
                for i in range(1,len(mcc_val)):
                    if "," not in str(mnc_val[i]) and "/" not in str(mnc_val[i]):
                        ind1=str(mcc_val[i]).rfind(".")
                        ind2=str(mnc_val[i]).rfind(".")
                        if ind1 != -1 and ind2 != -1:
                            val=str(mnc_val[i])[:ind2]
                            if len(val)==1:
                                val="0"+val
                            value=str(mcc_val[i])[:ind1]+val                  
                            sheet_wr.write(i,4,value)
                        else:
                            val1 = str(mcc_val[i])
                            val2 = str(mnc_val[i])
                            value = val1 + val2
                            sheet_wr.write(i,4,value)
                    else:
                        value=""
                        sheet_wr.write(i,4,value)
                        
            # """MNC and MNC individual columns are absent"""
            if mccmnc_absent==False and mcc_absent== True and mnc_absent==True:
                for i in range(1,len(mccmnc_val)):
                    value_mcc=str(mccmnc_val[i])[:3]
                    sheet_wr.write(i,2,value_mcc)
                    value_mnc=str(mccmnc_val[i])[3:]
                    if "." in value_mnc:
                        ind3=value_mnc.index(".")
                        value_mnc=value_mnc[:ind3]
                    sheet_wr.write(i,3,value_mnc)
    
            if not rate_present:
                move_to_day_folder(filename, edate, 'NoRates')
                return -1
    
            index = filename.rfind('.')
            filename1 = filename[:index]
            filename1 = filename1 + ' FORMATTED.xls'
            new_book.save(filename1)
            log_file().info("Source Compiler.log", "%s has been properly formatted." % filename)
            print("%s has been properly formatted." % filename)
            # os.remove(filename)
            return filename1
        except:
            error = sys.exc_info()[0]
            log_file().error("Source Compiler.log", "Error in format().excel_format: %s" % error)
            print("Error in format().excel_format: %s" % error)   
            #move_to_day_folder(filename, edate, '0BzlU44AWMToxOGtyYWZzSVAyNkE')

    # """Monty_is_special - formats the Rate to EUR, as it is not labeled properly """
    def monty_is_special(self, filename, og_file, edate):
        currency_rate = get_rate_for_date(edate)
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        og_book = xlrd.open_workbook(og_file)
        og_sheet = og_book.sheet_by_index(0)
        newbook = xlutils.copy.copy(book)
        sheet_wr = newbook.get_sheet(0)
        for x in range(sheet.nrows):
            if x > 0:
                curr = og_sheet.cell(x,10).value
                sheet_wr.write(x,6,curr)
                rate = sheet.cell(x,5).value
                for i in range(len(currency_list)):
                    if curr in currency_list[i]:
                        j = i
                        break

                converted = currency_rate[j]*float(rate)
                sheet_wr.write(x,7,converted)

        newbook.save(filename)
        return filename

    """ Quicksort sorts the a list and puts it in order, to_sort is a list of lists, index is the item in each list to use to sort"""
    def quicksort(self, to_sort, index):
        less = []
        equal = []
        greater = []

        if len(to_sort) > 1:
            pivot = to_sort[0]
            for x in to_sort:
                if x[index ]< pivot[index]:
                    less.append(x)
                if x[index] == pivot[index]:
                    equal.append(x)
                if x[index] > pivot[index]:
                    greater.append(x)

            return self.quicksort(less, index)+equal+self.quicksort(greater, index)

        else:
            return to_sort

    """ Parses through strings with multiple components and returns list with separated strings"""
    def separator(self, cell_val):
        c=0
        start=False 
        start_ind=0
        end_ind=0
        for i in range(len(cell_val)):
            k=cell_val[i].isalpha()
            m=cell_val[i].isspace()
            if i==0 and k==False and m==False:
                start=True
                start_ind=c+1
                c=c+1
            elif k==False and m==False:
                end_ind=c
                if start==False:
                    end_ind=c+1
                break
            c=c+1
        cell_val1=cell_val[end_ind:].strip()
        c1=end_ind
        start_ind1=c1
        for j in range(len(cell_val1)):
            k=cell_val[c1].isalpha()
            m=cell_val[i].isspace()
            if k==True and m==True:
                start_ind1=c1
            else:
                c1=c1+1       
        #country
        str1=cell_val[start_ind:end_ind-1]
        #operator
        str2=cell_val[start_ind1:].strip()
        return [str1,str2]

def file_clean(filename):
    index = filename.rfind('.')
    short = filename[:index]
    if os.path.isfile(short + '.xls'):
        os.remove(short + '.xls')

    if os.path.isfile(short + '.xlsx'):
        os.remove(short + '.xlsx')

    if os.path.isfile(short + '.csv'):
        os.remove(short + '.csv')

    if os.path.isfile(short + '.pdf'):
        os.remove(short + '.pdf')

    if os.path.isfile(short + ' FORMATTED.xls'):
        os.remove(short + ' FORMATTED.xls')

    if os.path.isfile(short + ' FORMATTED and FILTERED.xls'):
        os.remove(short + ' FORMATTED and FILTERED.xls')

    for i in range(10):
        renamed = short + '(' + str(i) + ')'
        if os.path.isfile(renamed + '.xls'):
            os.remove(renamed + '.xls')

        if os.path.isfile(renamed + '.xlsx'):
            os.remove(renamed + '.xlsx')

        if os.path.isfile(renamed + '.csv'):
            os.remove(renamed + '.csv')
        if os.path.isfile(short + '.pdf'):
            os.remove(short + '.pdf')

    log_file().info("Source Compiler.log", "All file versions of " + str(short) + " have been deleted.")
    print "All file versions of " + str(short) + " have been deleted."
##print(seperator(str1))
##print(seperator(str2))



"""-------------------------------------------------------------------------Main Code ------------------------------------------------------------"""
# format().excel_format('test.xls')
# print str(date.today())
# # date.timedelta(days=1)
# tomorrow = date.today() + timedelta(days=1)
