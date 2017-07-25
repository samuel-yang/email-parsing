from datetime import *
from fixerio import Fixerio
from Google_API_Manipulation import *
from write_log import *
import openpyxl
import os

global rate_in

base_currencies = ['EUR', 'GBP', 'CNY', 'MXN', 'AUD']
rate_in = ['USD']

def get_currency(currency, rate_in, date):
	fxrio = Fixerio(base=currency, symbols=rate_in, secure=True)
	return fxrio.historical_rates(date)

def get_rates(rate_in, date):
	currency_list = [float(1)]
	for currency in base_currencies:
		rate = get_currency(currency, rate_in, date)
		rate = rate['rates']['USD']
		currency_list.append(rate)

	return currency_list

def get_multipe_rates(dates):
	date_list = {}
	for date in dates:
		rate_list = get_rates(rate_in, date)
		date_list[date] = rate_list

	return date_list


# filename = 'Currency Exchange'
# if not os.path.isfile(filename+'.xlsx'):
# 	file_id = find_file_id(filename)
# 	export_sheet(file_id)


# book = openpyxl.load_workbook(filename+'.xlsx')
# sheet = book['Sheet1']
# date = date.today() - timedelta(days=60)
# rate_list = get_rates(rate_in, date)
# sheet.cell(row=2,column=1).value = date
# colnum = sheet.max_column
# for j in range(colnum-1):
# 	value = rate_list[j]
# 	sheet.cell(row=2, column=j+2).value = value

# book.save('Currency Exchange.xlsx')
# # print rate_list
# # print colnum

def get_rate_for_date(checkdate):
	filename = 'Currency Exchange'
	if not os.path.isfile(filename+'.xlsx'):
		file_id = find_file_id(filename)
		export_sheet(file_id)


	book = openpyxl.load_workbook(filename+'.xlsx')
	sheet = book['Sheet1']
	rownum = sheet.max_row
	colnum = sheet.max_column
	no_date = True
	rate_list = []
	temp_date = date.today()
	if checkdate > date.today():
		checkdate = date.today()

	for i in range(rownum-1):
		i = i + 1
		temp_date = sheet.cell(row=i+1,column=1).value
		if temp_date == None:
			temp_date = sheet.cell(row=i,column=1).value
			temp_date = temp_date.date()
			break		
		temp_date = temp_date.date() 
		if temp_date < checkdate:
			pass

		elif temp_date == checkdate:
			no_date = False
			rate_list = []
			for j in range(colnum-1):	
				value = sheet.cell(row=i+1,column=j+2).value
				rate_list.append(value)
			return rate_list
		
		row = i
			
	if no_date:
		temp_date = temp_date + timedelta(days=1)
		while temp_date <= checkdate:
			sheet.cell(row=row+2,column=1).value = temp_date
			rate_list = get_rates(rate_in, temp_date)
			for j in range(colnum-1):
				value = rate_list[j]
				sheet.cell(row=row+2,column=j+2).value = value
			temp_date = temp_date + timedelta(days=1)
			row = row + 1


	book.save('Currency Exchange.xlsx')
	return rate_list


# date = date.today() - timedelta(days=3)
# rate = get_rate_for_date(rate_in, date)
# print (date, rate)
# print ("all done")